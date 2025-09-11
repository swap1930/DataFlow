import pandas as pd
import os
import itertools
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
import plotly.express as px
import plotly.io as pio
from tempfile import NamedTemporaryFile
import datetime
import matplotlib.pyplot as plt

# Initialize Plotly Kaleido scope explicitly for reliable image export (e.g., on Render)
try:
    # These defaults can be tuned as needed
    if getattr(pio, "kaleido", None) and getattr(pio.kaleido, "scope", None):
        pio.kaleido.scope.default_format = "png"
        pio.kaleido.scope.default_width = 600
        pio.kaleido.scope.default_height = 400
        pio.kaleido.scope.scale = 1
        # Disable external dependencies for safety in server environments
        pio.kaleido.scope.mathjax = None
        pio.kaleido.scope.chromium_args = [
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
        ]
except Exception as _kaleido_init_err:
    print(f"⚠️ Kaleido initialization failed: {_kaleido_init_err}")

def process_excel_file(upload_dir, remove_fields="", number_of_relations=3,
                              description="", require_dashboard=True,
                              output_filename="processed_file.xlsx"):

    def make_json_serializable(data):
        """Recursively convert non-serializable types (like datetime) to strings."""
        if isinstance(data, list):
            return [make_json_serializable(d) for d in data]    
        elif isinstance(data, dict):
            return {k: make_json_serializable(v) for k, v in data.items()}
        elif isinstance(data, (datetime.datetime, datetime.date)):
            return data.isoformat()
        else:
            return data

    uploaded_files = os.listdir(upload_dir)
    if not uploaded_files:
        raise ValueError("No uploaded file found.")

    file_path = os.path.join(upload_dir, uploaded_files[0])
    ext = os.path.splitext(file_path)[1].lower()

    # Load file
    if ext in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path, engine='openpyxl')
    elif ext == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError("Unsupported file format.")

    # Clean data
    df.dropna(how='all', inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    df.replace('', pd.NA, inplace=True)
    df.dropna(inplace=True)

    # Remove specified fields
    if remove_fields:
        remove_list = [r.strip() for r in remove_fields.split(",")]
        df = df.drop(columns=remove_list, errors='ignore')

    # Identify categorical columns
    cat_fields = df.select_dtypes(include='object').nunique().sort_values(ascending=False).index.tolist()

    # Generate pivot tables
    pivot_tables = []
    pivot_data_for_frontend = []  # New: structured data for frontend
    
    if len(cat_fields) >= 2:
        combos = list(itertools.combinations(cat_fields, 2))[:number_of_relations]
        for col1, col2 in combos:
            pivot = pd.pivot_table(df, index=col1, columns=col2, aggfunc='size', fill_value=0)
            pivot_tables.append((f"{col1} vs {col2}", pivot))
            
            # Create structured data for frontend
            pivot_frontend = {
                "title": f"{col1} vs {col2}",
                "index_column": col1,
                "column_headers": [str(col) for col in pivot.columns],
                "data": []
            }
            
            # Convert pivot table to structured format
            for idx, row in pivot.iterrows():
                row_data = {"index": str(idx)}
                for col in pivot.columns:
                    row_data[str(col)] = int(row[col])
                pivot_frontend["data"].append(row_data)
            
            pivot_data_for_frontend.append(pivot_frontend)
            
    elif len(cat_fields) == 1:
        col = cat_fields[0]
        pivot = df[col].value_counts().to_frame(name='Count')
        pivot_tables.append((f"{col} Frequency", pivot))
        
        # Create structured data for frontend
        pivot_frontend = {
            "title": f"{col} Frequency",
            "index_column": col,
            "column_headers": ["Count"],
            "data": []
        }
        
        for idx, count in pivot.iterrows():
            pivot_frontend["data"].append({
                "index": str(idx),
                "Count": int(count['Count'])
            })
        
        pivot_data_for_frontend.append(pivot_frontend)
        
    else:
        numeric_cols = df.select_dtypes(include='number').columns.tolist()
        if numeric_cols:
            col = numeric_cols[0]
            pivot = df[col].value_counts().to_frame(name='Count')
            pivot_tables.append((f"{col} Frequency", pivot))
            
            # Create structured data for frontend
            pivot_frontend = {
                "title": f"{col} Frequency",
                "index_column": col,
                "column_headers": ["Count"],
                "data": []
            }
            
            for idx, count in pivot.iterrows():
                pivot_frontend["data"].append({
                    "index": str(idx),
                    "Count": int(count['Count'])
                })
            
            pivot_data_for_frontend.append(pivot_frontend)
        else:
            raise ValueError("No usable columns to generate relationships.")

    # Create workbook
    wb = Workbook()
    ws_cleaned = wb.active
    ws_cleaned.title = "CleanedData"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_cleaned.append(r)

    # Pivot tables sheet
    ws_pivot = wb.create_sheet("PivotTables")
    ws_pivot["A1"] = "📌 All Pivot Tables"
    ws_pivot["A1"].font = Font(bold=True, size=14)
    current_row = 3
    pivot_sheet_positions = []

    for title, pivot in pivot_tables:
        ws_pivot.merge_cells(start_row=current_row, start_column=1,
                             end_row=current_row, end_column=1 + len(pivot.columns))
        ws_pivot.cell(row=current_row, column=1).value = f"Pivot: {title}"
        ws_pivot.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        for r in dataframe_to_rows(pivot, index=True, header=True):
            for col_index, value in enumerate(r, start=1):
                ws_pivot.cell(row=current_row, column=col_index).value = value
            current_row += 1
        current_row += 2
        pivot_sheet_positions.append((title, pivot))

    # Dashboard: 1 chart per pivot table, 2 charts per row layout
    if require_dashboard and pivot_sheet_positions:
        ws_dash = wb.create_sheet("Dashboard")
        ws_dash.sheet_view.showGridLines = False
        ws_dash["B2"] = "📊 Dashboard - Auto Generated"
        ws_dash["B2"].font = Font(bold=True, size=14)

        chart_row_start = 5
        chart_spacing_row = 25
        chart_spacing_col = 10

        # Chart pool
        chart_types = ["bar", "pie", "line", "heatmap", "scatter", "area"]

        # Collect chart metadata and optional interactive JSON for frontend
        dashboard_charts_meta = []

        for idx, (title, pivot) in enumerate(pivot_sheet_positions):
            pivot_reset = pivot.reset_index()
            row_num = chart_row_start + (idx // 2) * chart_spacing_row
            col_num = 2 + (idx % 2) * chart_spacing_col

            chart_type = chart_types[idx % len(chart_types)]

            if chart_type == "bar":
                fig = px.bar(pivot_reset, x=pivot_reset.columns[0], y=pivot_reset.columns[1], title=title + " Bar Chart")
            elif chart_type == "pie":
                fig = px.pie(pivot_reset, names=pivot_reset.columns[0], values=pivot_reset.columns[1], title=title + " Pie Chart")
            elif chart_type == "line":
                fig = px.line(pivot_reset, x=pivot_reset.columns[0], y=pivot_reset.columns[1], title=title + " Line Chart")
            elif chart_type == "scatter":
                fig = px.scatter(pivot_reset, x=pivot_reset.columns[0], y=pivot_reset.columns[1], title=title + " Scatter Chart")
            elif chart_type == "area":
                fig = px.area(pivot_reset, x=pivot_reset.columns[0], y=pivot_reset.columns[1], title=title + " Area Chart")
            else:  # heatmap
                fig = px.imshow(pivot.values, labels=dict(x=pivot.columns.name, y=pivot.index.name, color="Count"),
                                x=pivot.columns, y=pivot.index, title=title + " Heatmap")

            try:
                with NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
                    # Export Plotly figure to PNG via Kaleido
                    pio.write_image(fig, tmpfile.name, width=600, height=400)
                    img = XLImage(tmpfile.name)
                img.anchor = f"{chr(64 + col_num)}{row_num}"
                ws_dash.add_image(img)
                # Save meta with Plotly JSON for frontend interactivity
                try:
                    dashboard_charts_meta.append({
                        "title": title,
                        "chart_type": chart_type,
                        "backend": "plotly",
                        "plotly_json": fig.to_json(),
                    })
                except Exception as _json_err:
                    print(f"⚠️ Plotly to_json failed for '{title}': {_json_err}")
            except Exception as _plotly_export_err:
                # Log the actual error and fall back to Matplotlib static image
                print(f"❌ Plotly image export failed for '{title}': {_plotly_export_err}")

                try:
                    # Simple Matplotlib fallback using the same pivot data
                    plt.close('all')
                    fig_mpl, ax = plt.subplots(figsize=(6, 4), dpi=100)

                    if chart_type in ("bar", "line", "area", "scatter") and pivot_reset.shape[1] >= 2:
                        x_values = pivot_reset.iloc[:, 0].astype(str)
                        y_values = pivot_reset.iloc[:, 1].astype(float)
                        if chart_type == "bar":
                            ax.bar(x_values, y_values, color="#3B82F6")
                        elif chart_type == "line":
                            ax.plot(x_values, y_values, color="#10B981", marker="o")
                        elif chart_type == "area":
                            ax.fill_between(range(len(y_values)), y_values, color="#6366F1", alpha=0.5)
                            ax.plot(range(len(y_values)), y_values, color="#6366F1")
                        else:  # scatter
                            ax.scatter(range(len(y_values)), y_values, color="#F59E0B")
                        ax.set_title(title)
                        ax.set_xlabel(str(pivot_reset.columns[0]))
                        ax.set_ylabel(str(pivot_reset.columns[1]))
                        ax.tick_params(axis='x', rotation=45)
                        fig_mpl.tight_layout()
                    elif chart_type == "pie" and pivot_reset.shape[1] >= 2:
                        sizes = pivot_reset.iloc[:, 1].astype(float)
                        labels = pivot_reset.iloc[:, 0].astype(str)
                        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
                        ax.axis('equal')
                        ax.set_title(title)
                        fig_mpl.tight_layout()
                    else:
                        # Generic table heat visualization
                        import numpy as _np
                        heat = pivot.values.astype(float) if hasattr(pivot, 'values') else _np.array([[0]])
                        im = ax.imshow(heat, cmap='Blues')
                        ax.set_title(title)
                        fig_mpl.colorbar(im, ax=ax)
                        fig_mpl.tight_layout()

                    with NamedTemporaryFile(suffix=".png", delete=False) as tmpfile2:
                        fig_mpl.savefig(tmpfile2.name, format="png")
                        img2 = XLImage(tmpfile2.name)
                    img2.anchor = f"{chr(64 + col_num)}{row_num}"
                    ws_dash.add_image(img2)

                    dashboard_charts_meta.append({
                        "title": title,
                        "chart_type": chart_type,
                        "backend": "matplotlib",
                        "plotly_json": None,
                    })
                except Exception as _mpl_err:
                    print(f"❌ Matplotlib fallback failed for '{title}': {_mpl_err}")

        # Attach dashboard meta for frontend use (interactive where available)
        ws_dash["B3"] = "(Metadata available in API response)"

    # Save file to temporary location (in memory)
    from io import BytesIO
    import base64
    
    # Create BytesIO buffer to save workbook in memory
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Convert to base64 for MongoDB storage
    file_content_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    # Prepare return data with JSON-serializable conversion
    processed_data = {
        "cleaned_data": make_json_serializable(df.to_dict(orient="records")),
        "pivot_tables": [],
        "has_dashboard": require_dashboard and bool(pivot_sheet_positions),
        "sheets": wb.sheetnames,
        "file_content_base64": file_content_base64,  # Store file content as base64
        "file_name": output_filename,
        "requested_relations": number_of_relations,
        "generated_relations": len(pivot_tables)
    }
    for title, pivot in pivot_tables:
        processed_data["pivot_tables"].append({
            "title": title,
            "index_column": pivot.index.name if pivot.index.name else "Index",
            "column_headers": [str(c) for c in pivot.columns],
            "data": make_json_serializable([
                {**{"index": str(idx)}, **{str(col): int(val) if pd.notna(val) else 0 for col, val in row.items()}}
                for idx, row in pivot.iterrows()
            ])
        })

    # Include dashboard chart JSONs if created
    try:
        if require_dashboard and 'ws_dash' in locals() and 'dashboard_charts_meta' in locals():
            processed_data["dashboard_charts"] = dashboard_charts_meta
    except Exception as _attach_meta_err:
        print(f"⚠️ Failed to attach dashboard chart metadata: {_attach_meta_err}")

    return processed_data
